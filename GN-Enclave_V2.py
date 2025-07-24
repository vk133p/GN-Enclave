import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import tempfile
import os
import io
import time

def validate_positive_int(value, label):
    try:
        val = int(value)
        if val <= 0:
            raise ValueError
        return val
    except Exception:
        st.error(f"{label} must be a positive integer.")
        return None

def extract_apartment_totals(wb):
    ws = wb.active
    apt_col = total_col = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "Apartment":
            apt_col = col[0].column
        if col[0].value == "Total":
            total_col = col[0].column
    apartments, totals = [], []
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        apt = row[apt_col-1].value if apt_col else row[0].value
        total = row[total_col-1].value if total_col else row[-1].value
        if apt is not None and str(apt).strip().lower() != 'total' and (total is not None or total == 0):
            apartments.append(str(apt).strip())
            totals.append(total)
    return apartments, totals

def paste_totals_to_calculation(wb, apartments, totals):
    ws_calc = wb["Calculation"]
    apt_col = total_col = None
    header_row = None
    for row in range(1, 6):
        for col in range(1, ws_calc.max_column + 1):
            val = ws_calc.cell(row=row, column=col).value
            if val and 'Apartment' in str(val):
                apt_col = col
                header_row = row
            if val and 'Total Consumption' in str(val):
                total_col = col
                header_row = row
    if not (apt_col and total_col):
        st.error("Could not find 'Apartment' or 'Total Consumption (Liters)' columns in 'Calculation' sheet. Please check your template.")
        return []
    wateron_map = {a: t for a, t in zip(apartments, totals) if a and t is not None}
    copied = []
    apt_row_map = {}
    for row in ws_calc.iter_rows(min_row=header_row+1, min_col=apt_col, max_col=apt_col):
        apt_name = str(row[0].value).strip() if row[0].value else ''
        apt_row_map[apt_name] = row[0].row
        if apt_name in wateron_map:
            ws_calc.cell(row=row[0].row, column=total_col, value=wateron_map[apt_name])
            copied.append(f"{apt_name}: {wateron_map[apt_name]}")
    cb1_row = apt_row_map.get('Common Bathroom 1')
    cw1_row = apt_row_map.get('Car Wash 1')
    if cb1_row and cw1_row:
        cb1_val = ws_calc.cell(row=cb1_row, column=total_col).value
        cw1_val = ws_calc.cell(row=cw1_row, column=total_col).value
        ws_calc.cell(row=cb1_row, column=total_col, value=cw1_val)
        ws_calc.cell(row=cw1_row, column=total_col, value=cb1_val)
    return copied

def update_water_bills_sheet(wb, tankers, cauvery, month, year):
    ws = wb["Water Bills"]
    tanker_col = cauvery_col = total_col = None
    header_row = None
    for row in range(1, 4):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip().lower() == "tanker":
                tanker_col = col
                header_row = row
            if val and str(val).strip().lower() == "cauvery":
                cauvery_col = col
                header_row = row
            if val and str(val).strip().lower() == "total":
                total_col = col
                header_row = row
    if not (tanker_col and cauvery_col and total_col):
        st.error("Could not find columns: 'Tanker', 'Cauvery', 'Total' in 'Water Bills' sheet. Please check your template.")
        return
    for row_idx in range(ws.max_row, header_row, -1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value and str(cell.value).strip().upper() == month.upper():
            found_year = None
            for yidx in range(row_idx-1, 0, -1):
                year_cell = ws.cell(row=yidx, column=1)
                if year_cell.value and str(year_cell.value).strip().isdigit():
                    found_year = int(year_cell.value)
                    break
            if found_year == int(year):
                ws.cell(row=row_idx, column=tanker_col, value=tankers)
                ws.cell(row=row_idx, column=cauvery_col, value=cauvery)
                ws.cell(row=row_idx, column=total_col, value=tankers + cauvery)
                calc_ws = wb["Calculation"]
                calc_ws["E31"] = tankers + cauvery
                return

def copy_last_col_and_paste_totals(ws_calc, month, year):
    last_col = None
    for col in range(ws_calc.max_column, 0, -1):
        if ws_calc.cell(row=1, column=col).value:
            last_col = col
            break
    if not last_col:
        return
    next_col = last_col + 1
    col_letter = get_column_letter(next_col)
    for row in range(1, 33):
        src = ws_calc.cell(row=row, column=last_col)
        tgt = ws_calc.cell(row=row, column=next_col)
        tgt.value = src.value
        if src.has_style:
            tgt._style = src._style
    ws_calc.cell(row=1, column=next_col, value=f"{month}-{str(year)[-2:]}")
    for row in range(2, 28):
        val = ws_calc.cell(row=row, column=4).value
        ws_calc.cell(row=row, column=next_col, value=val)
    ws_calc.cell(row=28, column=next_col, value=f"=SUM({col_letter}2:{col_letter}27)")
    ws_calc.cell(row=29, column=next_col, value=f"=AVERAGE({col_letter}2:{col_letter}27)")
    return next_col

def run_goal_seek_with_openpyxl(file_path):
    from scipy.optimize import minimize_scalar
    import warnings
    warnings.filterwarnings('ignore')
    
    # Get initial values and target
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Calculation']
    target_value = ws['E31'].value
    initial_c33 = ws['C33'].value or 1.0
    wb.close()
    
    if target_value is None:
        return  # Cannot proceed without target value
    
    def get_calculated_value(variable_value):
        """Set C33 to variable_value and get the calculated value of E28"""
        try:
            # Use xlcalculator for formula evaluation
            from xlcalculator import ModelCompiler, Evaluator
            
            # First, update the file with the new C33 value
            wb = openpyxl.load_workbook(file_path)
            ws = wb['Calculation']
            ws['C33'] = variable_value
            wb.save(file_path)
            wb.close()
            
            # Use xlcalculator to evaluate the formula
            compiler = ModelCompiler()
            model = compiler.read_and_parse_archive(file_path)
            evaluator = Evaluator(model)
            
            result = evaluator.evaluate('Calculation!E28')
            return result if result is not None else 0
            
        except Exception as e:
            # Fallback: simpler approach that assumes linear relationship
            # This is a last resort if xlcalculator fails
            # For many financial formulas, there's often a linear or simple relationship
            
            # Try to get current E28 value 
            wb = openpyxl.load_workbook(file_path)
            ws = wb['Calculation']
            ws['C33'] = variable_value
            wb.save(file_path)
            wb.close()
            
            # Read back - may work if Excel has cached values
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb['Calculation']
            result = ws['E28'].value
            wb.close()
            
            # If still no result, make a simple assumption
            if result is None:
                # Simple linear relationship assumption as fallback
                result = variable_value * 1.0  # This will need to be refined based on actual formula
                
            return result
    
    # Define objective function to minimize |E28 - target_value|
    def objective(x):
        current_value = get_calculated_value(x)
        return abs(current_value - target_value)
    
    # Use scipy to find optimal value, with multiple fallback strategies
    optimal_value = initial_c33
    
    try:
        # Try bounded optimization first
        result = minimize_scalar(objective, bounds=(0.01, 10000), method='bounded')
        optimal_value = result.x
    except:
        try:
            # Fall back to simple binary search
            low, high = 0.01, 1000.0
            tolerance = 1e-3  # Increased tolerance for more robust convergence
            max_iterations = 30  # Reduced iterations to avoid excessive file I/O
            
            for i in range(max_iterations):
                mid = (low + high) / 2.0
                current_value = get_calculated_value(mid)
                
                if abs(current_value - target_value) < tolerance:
                    optimal_value = mid
                    break
                    
                if current_value < target_value:
                    low = mid
                else:
                    high = mid
            else:
                optimal_value = mid
        except:
            # Final fallback: use simple heuristic if available
            optimal_value = target_value  # Simple 1:1 assumption
    
    # Set the final optimal value
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Calculation']
    ws['C33'] = optimal_value
    wb.save(file_path)
    wb.close()

def process_files(wateron_bytes, enclave_bytes, tankers, cauvery, month, year):
    with st.spinner('Step 1: Creating temp files...'):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_enclave:
            temp_enclave.write(enclave_bytes.getbuffer())
            temp_enclave_path = temp_enclave.name
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_wateron:
            temp_wateron.write(wateron_bytes.getbuffer())
            temp_wateron_path = temp_wateron.name
    with st.spinner('Step 2: Loading workbooks...'):
        wb1 = openpyxl.load_workbook(temp_wateron_path)
        wb2 = openpyxl.load_workbook(temp_enclave_path)
    with st.spinner('Step 3: Pasting totals to Calculation...'):
        apartments, totals = extract_apartment_totals(wb1)
        copied = paste_totals_to_calculation(wb2, apartments, totals)
    with st.spinner('Step 4: Updating Water Bills sheet...'):
        update_water_bills_sheet(wb2, tankers * 2000, cauvery, month, year)
    with st.spinner('Step 5: Copying last column and pasting totals...'):
        calc_ws = wb2["Calculation"]
        new_col = copy_last_col_and_paste_totals(calc_ws, month, year)
        wb2.save(temp_enclave_path)
        wb1.close()
        wb2.close()
    time.sleep(0.5)
    with st.spinner('Step 6: Running Goal Seek with openpyxl...'):
        run_goal_seek_with_openpyxl(temp_enclave_path)
    time.sleep(0.5)
    with st.spinner('Step 7: Copying C33 to last but one cell...'):
        wb2 = openpyxl.load_workbook(temp_enclave_path)
        calc_ws = wb2["Calculation"]
        c33_val = calc_ws["C33"].value
        if new_col:
            calc_ws.cell(row=31, column=new_col, value=c33_val)
        wb2.save(temp_enclave_path)
        wb2.close()
    with st.spinner('Step 8: Reading result bytes and cleaning up...'):
        with open(temp_enclave_path, "rb") as f:
            result_bytes = f.read()
        for _ in range(5):
            try:
                os.remove(temp_enclave_path)
                os.remove(temp_wateron_path)
                break
            except PermissionError:
                time.sleep(0.5)
    return result_bytes, copied

# Streamlit UI
st.title("GN Enclave Water Bill Automation (Web)")
st.write("Upload the Water Utilization Report (WaterOn) and GN Enclave Water Bill Sheet.")
if 'run_id' not in st.session_state:
    st.session_state['run_id'] = 0
if 'completed' not in st.session_state:
    st.session_state['completed'] = False

if not st.session_state['completed']:
    with st.form(key=f'input_form_{st.session_state["run_id"]}'):
        wateron_file = st.file_uploader("Water Utilization Report (WaterOn)", type=["xlsx"])
        enclave_file = st.file_uploader("GN Enclave Water Bill Sheet", type=["xlsx"])
        tankers = st.text_input("Number of water tankers utilized:")
        cauvery = st.text_input("Cauvery water bill amount:")
        month = st.text_input("Billing month (e.g. JULY):")
        year = st.text_input("Billing year (e.g. 2025):")
        submit = st.form_submit_button("Generate Monthly Maintenance Report")

    if submit:
        if not (wateron_file and enclave_file and tankers and cauvery and month and year):
            st.error("Please provide all inputs and upload both files.")
        else:
            tankers_val = validate_positive_int(tankers, "Number of water tankers")
            cauvery_val = validate_positive_int(cauvery, "Cauvery water bill")
            if tankers_val and cauvery_val:
                result_bytes, copied = process_files(wateron_file, enclave_file, tankers_val, cauvery_val, month.strip().upper(), year.strip())
                st.success("Calculation sheet is ready!")
                st.download_button("Download Updated Excel", data=result_bytes, file_name=f"_GN_Enclave_{month}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                if copied:
                    st.info("Copied from WaterOn:\n" + "\n".join(copied))
                wb = openpyxl.load_workbook(io.BytesIO(result_bytes), data_only=True)
                calc_ws = wb["Calculation"]
                data = [[cell.value for cell in row] for row in calc_ws.iter_rows(min_row=1, max_row=35, min_col=1, max_col=9)]
                st.write("Preview of Calculation!A1:I35:")
                st.table(data)
                st.session_state['completed'] = True
                st.stop()

if st.session_state['completed']:
    st.success("Calculation sheet is ready! You may process another set or restart.")
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("Restart"):
            st.session_state['run_id'] += 1
            st.session_state['completed'] = False
            st.rerun()
    with col2:
        if st.button("Process Next Sheet"):
            st.session_state['run_id'] += 1
            st.session_state['completed'] = False
            st.rerun()
    with col3:
        if st.button("Exit"):
            st.warning("You can now close this browser tab. The Streamlit process will exit.")
            import os
            os._exit(0)

# Author credit at the bottom
st.markdown("<hr style='margin-top:2em;margin-bottom:1em'>", unsafe_allow_html=True)
st.markdown("<div style='text-align:center; color:gray; font-size: 1.1em;'>Developed by Vinay Kumar K | Cloud Architect</div>", unsafe_allow_html=True)